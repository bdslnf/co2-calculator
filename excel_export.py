"""
Excel-Export für Management-Reports
Erstellt professionelle Excel-Dateien mit mehreren Sheets
"""

from pathlib import Path
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def formatiere_waehrung(ws, zeile_start, zeile_end, spalte):
    """Formatiert Spalte als Währung."""
    for row in range(zeile_start, zeile_end + 1):
        cell = ws.cell(row=row, column=spalte)
        cell.number_format = '#,##0 "CHF"'


def formatiere_prozent(ws, zeile_start, zeile_end, spalte):
    """Formatiert Spalte als Prozent."""
    for row in range(zeile_start, zeile_end + 1):
        cell = ws.cell(row=row, column=spalte)
        cell.number_format = '0.0"%"'


def formatiere_header(ws, zeile):
    """Formatiert Header-Zeile."""
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[zeile]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def erstelle_uebersicht_sheet(
    wb: Workbook,
    portfolio_stats: Dict,
    gebaeude_df: pd.DataFrame
) -> None:
    """
    Erstellt Übersichts-Sheet.
    
    Args:
        wb: Workbook-Objekt
        portfolio_stats: Portfolio-Statistiken
        gebaeude_df: DataFrame mit Gebäuden
    """
    ws = wb.create_sheet("Übersicht", 0)
    
    # Titel
    ws["A1"] = "CO₂ NEUTRALITY PATH CALCULATOR"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:D1")
    
    ws["A2"] = "Executive Summary - Portfolio-Analyse"
    ws["A2"].font = Font(size=12, italic=True)
    ws.merge_cells("A2:D2")
    
    # Kennzahlen
    row = 4
    ws[f"A{row}"] = "PORTFOLIO-KENNZAHLEN"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    
    row += 1
    kennzahlen = [
        ("Anzahl Gebäude:", portfolio_stats.get("anzahl_gebaeude", 0)),
        ("Gesamt-Emissionen [t CO₂e/Jahr]:", portfolio_stats.get("gesamt_emissionen_t_jahr", 0)),
        ("Ø Emissionen pro Gebäude [t/Jahr]:", portfolio_stats.get("durchschnitt_emissionen_t_jahr", 0)),
    ]
    
    if portfolio_stats.get("gesamt_flaeche_m2"):
        kennzahlen.extend([
            ("Gesamt-Fläche [m²]:", portfolio_stats["gesamt_flaeche_m2"]),
            ("Ø Emissionen [kg CO₂e/m²]:", portfolio_stats.get("durchschnitt_emissionen_kg_m2", 0)),
        ])
    
    for label, wert in kennzahlen:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = wert
        ws[f"B{row}"].number_format = "#,##0.0"
        row += 1
    
    # Heizungstypen
    row += 2
    ws[f"A{row}"] = "HEIZUNGSTYPEN-VERTEILUNG"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    ws[f"A{row}"] = "Heizungstyp"
    ws[f"B{row}"] = "Anzahl"
    ws[f"C{row}"] = "Anteil [%]"
    formatiere_header(ws, row)
    row += 1
    
    for typ, anzahl in portfolio_stats.get("heizungstypen_verteilung", {}).items():
        prozent = (anzahl / portfolio_stats["anzahl_gebaeude"]) * 100 if portfolio_stats["anzahl_gebaeude"] > 0 else 0
        ws[f"A{row}"] = typ
        ws[f"B{row}"] = anzahl
        ws[f"C{row}"] = prozent / 100
        ws[f"C{row}"].number_format = "0.0%"
        row += 1
    
    # Spaltenbreiten
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15


def erstelle_sanierungen_sheet(
    wb: Workbook,
    sanierungen_df: pd.DataFrame
) -> None:
    """
    Erstellt Sanierungs-Empfehlungen Sheet.
    
    Args:
        wb: Workbook-Objekt
        sanierungen_df: Priorisierter DataFrame mit Sanierungen
    """
    ws = wb.create_sheet("Sanierungsempfehlungen")
    
    # Titel
    ws["A1"] = "SANIERUNGSEMPFEHLUNGEN"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:H1")
    
    # Spalten auswählen
    spalten = [
        "rang", "name", "investition_netto_chf", "co2_einsparung_kg_jahr",
        "amortisation_jahre", "roi_prozent", "npv_chf", "prioritaets_score"
    ]
    
    # Nur existierende Spalten
    spalten = [s for s in spalten if s in sanierungen_df.columns]
    
    df_export = sanierungen_df[spalten].copy()
    
    # Umbenennen
    df_export.columns = [
        "Rang", "Massnahme", "Investition [CHF]", "CO₂-Reduktion [kg/Jahr]",
        "Amortisation [Jahre]", "ROI [%]", "NPV [CHF]", "Score"
    ]
    
    # CO₂ in Tonnen
    if "CO₂-Reduktion [kg/Jahr]" in df_export.columns:
        df_export["CO₂-Reduktion [t/Jahr]"] = df_export["CO₂-Reduktion [kg/Jahr]"] / 1000
        df_export = df_export.drop("CO₂-Reduktion [kg/Jahr]", axis=1)
    
    # DataFrame einfügen
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Header formatieren
    formatiere_header(ws, 3)
    
    # Währungsformatierung (Spalte C, E, G)
    if len(df_export) > 0:
        formatiere_waehrung(ws, 4, 3 + len(df_export), 3)  # Investition
        formatiere_waehrung(ws, 4, 3 + len(df_export), 7)  # NPV
    
    # Spaltenbreiten
    ws.column_dimensions["B"].width = 35  # Massnahme
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18


def erstelle_wirtschaftlichkeit_sheet(
    wb: Workbook,
    sanierung: Dict
) -> None:
    """
    Erstellt detailliertes Wirtschaftlichkeits-Sheet für Top-Sanierung.
    
    Args:
        wb: Workbook-Objekt
        sanierung: Sanierung mit Wirtschaftlichkeitsdaten
    """
    ws = wb.create_sheet("Wirtschaftlichkeit_Detail")
    
    # Titel
    ws["A1"] = f"WIRTSCHAFTLICHKEITSANALYSE: {sanierung.get('name', 'N/A')}"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:D1")
    
    row = 3
    
    # Investition
    ws[f"A{row}"] = "INVESTITION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    inv_daten = [
        ("Brutto-Investition:", sanierung.get("investition_brutto_chf", 0)),
        ("Fördergelder:", sanierung.get("foerderung_chf", 0)),
        ("Netto-Investition:", sanierung.get("investition_netto_chf", 0)),
    ]
    
    for label, wert in inv_daten:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = wert
        ws[f"B{row}"].number_format = '#,##0 "CHF"'
        row += 1
    
    row += 1
    
    # Einsparungen
    ws[f"A{row}"] = "JÄHRLICHE EINSPARUNGEN"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    einspar_daten = [
        ("Energiekosten:", sanierung.get("energiekosteneinsparung_chf", 0)),
        ("CO₂-Abgaben:", sanierung.get("co2_abgabe_einsparung_chf", 0)),
        ("Gesamt:", sanierung.get("gesamteinsparung_chf_jahr", 0)),
    ]
    
    for label, wert in einspar_daten:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = wert
        ws[f"B{row}"].number_format = '#,##0 "CHF"'
        row += 1
    
    row += 1
    
    # KPIs
    ws[f"A{row}"] = "KENNZAHLEN"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    
    kpi_daten = [
        ("Amortisationszeit:", sanierung.get("amortisation_jahre", 0), "Jahre"),
        ("ROI (über Lebensdauer):", sanierung.get("roi_prozent", 0), "%"),
        ("Nettobarwert (NPV):", sanierung.get("npv_chf", 0), "CHF"),
        ("CO₂-Reduktion:", sanierung.get("co2_einsparung_kg_jahr", 0) / 1000, "t/Jahr"),
    ]
    
    for label, wert, einheit in kpi_daten:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = wert
        ws[f"C{row}"] = einheit
        if einheit == "CHF":
            ws[f"B{row}"].number_format = '#,##0'
        else:
            ws[f"B{row}"].number_format = "#,##0.0"
        row += 1
    
    # Cashflow-Tabelle
    if "cashflow_tabelle" in sanierung:
        row += 2
        ws[f"A{row}"] = "CASHFLOW-VERLAUF"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        cf_df = sanierung["cashflow_tabelle"]
        
        for r_idx, cf_row in enumerate(dataframe_to_rows(cf_df, index=False, header=True), row):
            for c_idx, value in enumerate(cf_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        formatiere_header(ws, row)
        formatiere_waehrung(ws, row + 1, row + len(cf_df), 2)
        formatiere_waehrung(ws, row + 1, row + len(cf_df), 3)
    
    # Spaltenbreiten
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12


def exportiere_portfolio_excel(
    output_path: Path,
    portfolio_stats: Dict,
    gebaeude_df: pd.DataFrame,
    sanierungen_df: pd.DataFrame = None,
    top_sanierung: Dict = None
) -> Path:
    """
    Exportiert vollständigen Portfolio-Report als Excel.
    
    Args:
        output_path: Pfad zur Excel-Datei
        portfolio_stats: Portfolio-Statistiken
        gebaeude_df: DataFrame mit Gebäuden
        sanierungen_df: Optional, priorisierte Sanierungen
        top_sanierung: Optional, detaillierte Top-Sanierung
        
    Returns:
        Pfad zur erstellten Datei
    """
    wb = Workbook()
    
    # Default-Sheet entfernen
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Übersicht
    erstelle_uebersicht_sheet(wb, portfolio_stats, gebaeude_df)
    
    # Gebäude-Daten
    ws_gebaeude = wb.create_sheet("Gebäude-Daten")
    
    # Spalten auswählen
    geb_spalten = ["gebaeude_id", "heizung_typ", "jahresverbrauch_kwh", 
                   "strom_kwh_jahr", "emissionen_gesamt_t"]
    if "flaeche_m2" in gebaeude_df.columns:
        geb_spalten.insert(1, "flaeche_m2")
    if "baujahr" in gebaeude_df.columns:
        geb_spalten.insert(1, "baujahr")
    
    geb_spalten = [s for s in geb_spalten if s in gebaeude_df.columns]
    df_export = gebaeude_df[geb_spalten].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_gebaeude.cell(row=r_idx, column=c_idx, value=value)
    
    formatiere_header(ws_gebaeude, 1)
    
    # Sanierungen
    if sanierungen_df is not None and len(sanierungen_df) > 0:
        erstelle_sanierungen_sheet(wb, sanierungen_df)
    
    # Wirtschaftlichkeit
    if top_sanierung is not None:
        erstelle_wirtschaftlichkeit_sheet(wb, top_sanierung)
    
    # Speichern
    wb.save(output_path)
    
    return output_path


def exportiere_empfehlungsbericht(
    output_path: Path,
    empfehlungstext: str
) -> Path:
    """
    Exportiert Empfehlungsbericht als einfache Text-Datei.
    
    Args:
        output_path: Pfad zur Datei
        empfehlungstext: Formatierter Empfehlungstext
        
    Returns:
        Pfad zur Datei
    """
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(empfehlungstext)
    
    return output_path
